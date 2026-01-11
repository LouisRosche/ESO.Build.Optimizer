#!/usr/bin/env python3
"""
ESO Feature Documentation - Excel Generator

Compiles all phase JSON data into a single normalized Excel dataset.
"""

import json
import os
from pathlib import Path
from typing import Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Schema columns
COLUMNS = [
    'feature_id', 'system', 'category', 'subcategory', 'feature_type',
    'name', 'parent_feature', 'class_restriction', 'unlock_method',
    'resource_type', 'resource_cost', 'cast_time', 'target_type',
    'range_m', 'radius_m', 'duration_sec', 'cooldown_sec', 'base_effect',
    'scaling_stat', 'max_ranks', 'rank_progression', 'stages',
    'points_per_stage', 'compatible_grimoires', 'buff_debuff_granted',
    'synergy', 'tags', 'dlc_required', 'patch_updated', 'source_url'
]

def normalize_feature(feature: dict) -> dict:
    """Normalize field names to match expected schema."""
    # Field name mappings
    mappings = {
        'feature_name': 'name',
        'skill_name': 'name',
        'description': 'base_effect',
        'effect': 'base_effect',
        'cost': 'resource_cost',
        'target': 'target_type',
        'range': 'range_m',
        'radius': 'radius_m',
        'duration': 'duration_sec',
        'cooldown': 'cooldown_sec',
        'unlock_requirements': 'unlock_method',
        'morph_of': 'parent_feature',
        'class': 'class_restriction',
    }

    normalized = {}
    for key, value in feature.items():
        new_key = mappings.get(key, key)
        normalized[new_key] = value

    # Normalize feature_type values
    if 'feature_type' in normalized:
        ft = str(normalized['feature_type']).upper()
        if ft in ['ACTIVE', 'SKILL']:
            normalized['feature_type'] = 'ACTIVE'
        elif ft in ['ULT', 'ULTIMATE']:
            normalized['feature_type'] = 'ULTIMATE'

    return normalized


def load_phase_data(raw_dir: Path) -> list[dict]:
    """Load all phase JSON files and combine into single list."""
    all_features = []

    # Find all JSON files matching phase patterns
    json_files = sorted(raw_dir.glob('phase*.json'))

    for file_path in json_files:
        print(f"Loading {file_path.name}...")
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                features = []
                if isinstance(data, list):
                    features = data
                elif isinstance(data, dict) and 'features' in data:
                    features = data['features']

                # Normalize each feature
                normalized = [normalize_feature(f) for f in features]
                all_features.extend(normalized)
                print(f"  -> {len(features)} features")
        except json.JSONDecodeError as e:
            print(f"  ERROR: Invalid JSON in {file_path.name}: {e}")

    return all_features


def validate_data(features: list[dict]) -> tuple[list[str], list[str]]:
    """Validate data integrity and return errors/warnings."""
    errors = []
    warnings = []
    feature_ids = set()
    parent_ids = set()

    # First pass - collect all IDs
    for f in features:
        fid = f.get('feature_id')
        if fid:
            if fid in feature_ids:
                errors.append(f"Duplicate feature_id: {fid}")
            feature_ids.add(fid)

    # Second pass - validate references
    for f in features:
        fid = f.get('feature_id', 'UNKNOWN')

        # Required fields
        for req in ['feature_id', 'system', 'category', 'name']:
            if not f.get(req):
                errors.append(f"{fid}: Missing required field '{req}'")

        # Morph validation
        ftype = f.get('feature_type', '')
        if ftype in ['MORPH_A', 'MORPH_B']:
            parent = f.get('parent_feature')
            if not parent:
                errors.append(f"{fid}: Morph missing parent_feature")
            elif parent not in feature_ids:
                warnings.append(f"{fid}: parent_feature '{parent}' not found")

        # Script validation
        if ftype in ['FOCUS_SCRIPT', 'SIGNATURE_SCRIPT', 'AFFIX_SCRIPT']:
            if not f.get('compatible_grimoires'):
                errors.append(f"{fid}: Script missing compatible_grimoires")

    return errors, warnings


def create_excel(features: list[dict], output_path: Path):
    """Create formatted Excel workbook from features."""
    # Create DataFrame
    df = pd.DataFrame(features)

    # Ensure all columns exist in correct order
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = None
    df = df[COLUMNS]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "ESO Features"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Handle None/NaN/lists
            if value is None:
                cell.value = None
            elif isinstance(value, (list, dict)):
                cell.value = json.dumps(value)
            elif isinstance(value, float) and pd.isna(value):
                cell.value = None
            else:
                cell.value = str(value) if not isinstance(value, (int, float, str, bool)) else value
            cell.alignment = cell_align
            cell.border = thin_border

    # Column widths
    column_widths = {
        'feature_id': 35, 'system': 12, 'category': 12, 'subcategory': 20,
        'feature_type': 15, 'name': 30, 'parent_feature': 35, 'class_restriction': 15,
        'unlock_method': 25, 'resource_type': 12, 'resource_cost': 12,
        'cast_time': 15, 'target_type': 12, 'range_m': 10, 'radius_m': 10,
        'duration_sec': 12, 'cooldown_sec': 12, 'base_effect': 50,
        'scaling_stat': 20, 'max_ranks': 10, 'rank_progression': 40,
        'stages': 8, 'points_per_stage': 15, 'compatible_grimoires': 30,
        'buff_debuff_granted': 30, 'synergy': 20, 'tags': 35,
        'dlc_required': 15, 'patch_updated': 12, 'source_url': 45
    }

    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = column_widths.get(col_name, 15)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Save
    wb.save(output_path)
    print(f"\nExcel file saved to: {output_path}")
    print(f"Total rows: {len(df)}")


def main():
    """Main entry point."""
    base_dir = Path(__file__).parent.parent
    raw_dir = base_dir / 'data' / 'raw'
    output_dir = base_dir / 'data' / 'compiled'
    output_path = output_dir / 'eso_features_complete.xlsx'

    # Ensure output directory exists
    output_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("ESO Feature Documentation - Excel Generator")
    print("=" * 60)

    # Load all phase data
    print("\nLoading phase data...")
    features = load_phase_data(raw_dir)
    print(f"\nTotal features loaded: {len(features)}")

    if not features:
        print("ERROR: No features found. Ensure JSON files exist in data/raw/")
        return

    # Validate
    print("\nValidating data...")
    errors, warnings = validate_data(features)

    if warnings:
        print(f"\nWarnings ({len(warnings)}):")
        for w in warnings[:10]:
            print(f"  - {w}")
        if len(warnings) > 10:
            print(f"  ... and {len(warnings) - 10} more")

    if errors:
        print(f"\nErrors ({len(errors)}):")
        for e in errors[:10]:
            print(f"  - {e}")
        if len(errors) > 10:
            print(f"  ... and {len(errors) - 10} more")
        print("\nProceeding with Excel generation despite errors...")
    else:
        print("  -> All validations passed!")

    # Generate Excel
    print("\nGenerating Excel file...")
    create_excel(features, output_path)

    print("\n" + "=" * 60)
    print("Generation complete!")
    print("=" * 60)


if __name__ == '__main__':
    main()
